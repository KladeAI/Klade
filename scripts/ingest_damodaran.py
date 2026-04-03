#!/usr/bin/env python3
"""Ingest Damodaran Industry Data into Supabase."""
import json, os, time, uuid, requests, io, tempfile
from datetime import datetime, timezone

# Config
ENV_PATH = "/home/Arjun/.openclaw/workspace/startup/klade-analyst/.env"
def load_env():
    env = {}
    with open(ENV_PATH) as f:
        for line in f:
            if '=' in line and not line.startswith('#'):
                k, v = line.strip().split('=', 1)
                env[k] = v
    return env

env = load_env()
SUPABASE_URL = env["SUPABASE_URL"]
SUPABASE_KEY = env["SUPABASE_SERVICE_ROLE_KEY"]
HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=minimal"
}

YEAR = 2026  # Current dataset year

# Damodaran Excel files to download
DATASETS = {
    "betas_us": {
        "url": "https://pages.stern.nyu.edu/~adamodar/pc/datasets/betas.xls",
        "region": "US",
        "metrics": "betas"
    },
    "betas_europe": {
        "url": "https://pages.stern.nyu.edu/~adamodar/pc/datasets/betaEurope.xls",
        "region": "Europe",
        "metrics": "betas"
    },
    "betas_global": {
        "url": "https://pages.stern.nyu.edu/~adamodar/pc/datasets/betaGlobal.xls",
        "region": "Global",
        "metrics": "betas"
    },
    "betas_emerging": {
        "url": "https://pages.stern.nyu.edu/~adamodar/pc/datasets/betaemerg.xls",
        "region": "Emerging Markets",
        "metrics": "betas"
    },
    "country_risk_premium": {
        "url": "https://pages.stern.nyu.edu/~adamodar/pc/datasets/ctryprem.xlsx",
        "region": "Global",
        "metrics": "country_risk_premium"
    },
}

# Additional US-focused datasets (CSV from HTML pages)
US_DATASETS_HTML = {
    "wacc": "https://pages.stern.nyu.edu/~adamodar/New_Home_Page/datafile/wacc.html",
    "EVMultiples": "https://pages.stern.nyu.edu/~adamodar/New_Home_Page/datafile/vebitda.html",
    "margins": "https://pages.stern.nyu.edu/~adamodar/New_Home_Page/datafile/margin.html",
}

import openpyxl
import xlrd

def download_file(url):
    """Download a file and return bytes."""
    resp = requests.get(url, headers={"User-Agent": "Klade AI dev@kladeai.com"}, timeout=30)
    resp.raise_for_status()
    return resp.content

def parse_xls_betas(data, region, source_url):
    """Parse Damodaran beta XLS files."""
    rows = []
    try:
        wb = xlrd.open_workbook(file_contents=data)
        for sheet_name in wb.sheet_names():
            ws = wb.sheet_by_name(sheet_name)
            if ws.nrows < 2:
                continue
            # Find header row
            header_row = None
            for r in range(min(20, ws.nrows)):
                cell_val = str(ws.cell_value(r, 0)).strip().lower()
                if 'industry' in cell_val or 'name' in cell_val:
                    header_row = r
                    break
            if header_row is None:
                # Try first row with content
                header_row = 0
            
            headers = [str(ws.cell_value(header_row, c)).strip() for c in range(ws.ncols)]
            
            for r in range(header_row + 1, ws.nrows):
                industry = str(ws.cell_value(r, 0)).strip()
                if not industry or industry.lower() in ('total', 'total market', ''):
                    continue
                
                for c in range(1, ws.ncols):
                    val = ws.cell_value(r, c)
                    metric_name = headers[c] if c < len(headers) else f"col_{c}"
                    if not metric_name or metric_name.lower() in ('', 'industry name'):
                        continue
                    try:
                        numeric_val = float(val) if val != '' else None
                    except (ValueError, TypeError):
                        continue
                    if numeric_val is not None:
                        rows.append({
                            "id": str(uuid.uuid4()),
                            "industry": industry,
                            "metric": metric_name,
                            "region": region,
                            "year": YEAR,
                            "value": numeric_val,
                            "source_url": source_url,
                            "created_at": datetime.now(timezone.utc).isoformat()
                        })
    except Exception as e:
        print(f"  Error parsing XLS: {e}")
    return rows

def parse_xlsx_country_risk(data, source_url):
    """Parse country risk premium XLSX."""
    rows = []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        for ws in wb.worksheets:
            header_row = None
            headers = []
            for r_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=False), 1):
                cell_vals = [str(c.value).strip() if c.value else '' for c in row]
                if any('country' in v.lower() for v in cell_vals[:3]):
                    header_row = r_idx
                    headers = cell_vals
                    break
            
            if header_row is None:
                continue
            
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                vals = list(row)
                country = str(vals[0]).strip() if vals[0] else ''
                if not country or country.lower() in ('none', 'total', ''):
                    continue
                
                for c_idx in range(1, len(vals)):
                    metric_name = headers[c_idx] if c_idx < len(headers) else f"col_{c_idx}"
                    if not metric_name:
                        continue
                    try:
                        numeric_val = float(vals[c_idx]) if vals[c_idx] is not None else None
                    except (ValueError, TypeError):
                        continue
                    if numeric_val is not None:
                        rows.append({
                            "id": str(uuid.uuid4()),
                            "industry": country,
                            "metric": f"country_risk_{metric_name}",
                            "region": country,
                            "year": YEAR,
                            "value": numeric_val,
                            "source_url": source_url,
                            "created_at": datetime.now(timezone.utc).isoformat()
                        })
    except Exception as e:
        print(f"  Error parsing XLSX: {e}")
    return rows

def parse_html_table(url, metric_prefix):
    """Parse Damodaran HTML tables."""
    rows = []
    try:
        resp = requests.get(url, headers={"User-Agent": "Klade AI dev@kladeai.com"}, timeout=30)
        resp.raise_for_status()
        html = resp.text
        
        # Simple HTML table parser
        import re
        # Find all tables
        tables = re.findall(r'<table[^>]*>(.*?)</table>', html, re.DOTALL | re.IGNORECASE)
        for table in tables:
            # Extract rows
            tr_pattern = re.findall(r'<tr[^>]*>(.*?)</tr>', table, re.DOTALL | re.IGNORECASE)
            if len(tr_pattern) < 2:
                continue
            
            # Get headers
            header_cells = re.findall(r'<t[hd][^>]*>(.*?)</t[hd]>', tr_pattern[0], re.DOTALL | re.IGNORECASE)
            headers = [re.sub(r'<[^>]+>', '', h).strip() for h in header_cells]
            
            if not headers or not any('industry' in h.lower() for h in headers[:2]):
                continue
            
            for tr in tr_pattern[1:]:
                cells = re.findall(r'<td[^>]*>(.*?)</td>', tr, re.DOTALL | re.IGNORECASE)
                cell_texts = [re.sub(r'<[^>]+>', '', c).strip() for c in cells]
                
                if not cell_texts or not cell_texts[0]:
                    continue
                
                industry = cell_texts[0]
                if industry.lower() in ('total', 'total market', ''):
                    continue
                
                for c_idx in range(1, len(cell_texts)):
                    metric_name = headers[c_idx] if c_idx < len(headers) else f"col_{c_idx}"
                    if not metric_name:
                        continue
                    # Clean numeric value
                    val_str = cell_texts[c_idx].replace(',', '').replace('%', '').replace('$', '').replace('NA', '').strip()
                    try:
                        numeric_val = float(val_str) if val_str else None
                    except (ValueError, TypeError):
                        continue
                    if numeric_val is not None:
                        rows.append({
                            "id": str(uuid.uuid4()),
                            "industry": industry,
                            "metric": f"{metric_prefix}_{metric_name}",
                            "region": "US",
                            "year": YEAR,
                            "value": numeric_val,
                            "source_url": url,
                            "created_at": datetime.now(timezone.utc).isoformat()
                        })
    except Exception as e:
        print(f"  Error parsing HTML {url}: {e}")
    return rows

def batch_insert(rows, table="damodaran_industry"):
    """Insert rows in batches with 2-second delays."""
    BATCH_SIZE = 500
    inserted = 0
    errors = 0
    for i in range(0, len(rows), BATCH_SIZE):
        batch = rows[i:i+BATCH_SIZE]
        r = requests.post(
            f"{SUPABASE_URL}/rest/v1/{table}",
            headers=HEADERS,
            json=batch
        )
        if r.status_code in (200, 201):
            inserted += len(batch)
            print(f"    Inserted batch {i//BATCH_SIZE + 1}: {len(batch)} rows (total: {inserted})")
        else:
            errors += len(batch)
            print(f"    ERROR batch {i//BATCH_SIZE + 1}: {r.status_code} - {r.text[:200]}")
        if i + BATCH_SIZE < len(rows):
            time.sleep(2)
    return inserted, errors

# Main execution
total_inserted = 0
total_errors = 0

# 1. Process Excel datasets
for name, config in DATASETS.items():
    print(f"\n--- Processing {name} ---")
    try:
        data = download_file(config["url"])
        print(f"  Downloaded {len(data)} bytes")
        
        if config["url"].endswith('.xlsx'):
            rows = parse_xlsx_country_risk(data, config["url"])
        else:
            rows = parse_xls_betas(data, config["region"], config["url"])
        
        print(f"  Parsed {len(rows)} rows")
        if rows:
            ins, err = batch_insert(rows)
            total_inserted += ins
            total_errors += err
    except Exception as e:
        print(f"  FAILED: {e}")

# 2. Process HTML datasets
for name, url in US_DATASETS_HTML.items():
    print(f"\n--- Processing HTML: {name} ---")
    try:
        rows = parse_html_table(url, name)
        print(f"  Parsed {len(rows)} rows")
        if rows:
            ins, err = batch_insert(rows)
            total_inserted += ins
            total_errors += err
    except Exception as e:
        print(f"  FAILED: {e}")

# Verify
r = requests.get(
    f"{SUPABASE_URL}/rest/v1/damodaran_industry?select=id",
    headers={**HEADERS, "Prefer": "count=exact", "Range": "0-0"}
)
count = r.headers.get("content-range", "unknown")
print(f"\n=== Damodaran Industry Data Summary ===")
print(f"Total inserted: {total_inserted}")
print(f"Total errors: {total_errors}")
print(f"DB count: {count}")
